import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
import xlsxwriter
from io import BytesIO, StringIO
from typing import List, Dict, Any, Union
import csv
import logging

logger = logging.getLogger(__name__)

class FileProcessor:
    """Handles reading and writing CSV and Excel files"""
    
    @staticmethod
    def read_file(file_content: bytes, file_extension: str) -> pd.DataFrame:
        """Read CSV or Excel file and return DataFrame"""
        try:
            if file_extension in ['.xlsx', '.xls']:
                return pd.read_excel(BytesIO(file_content), sheet_name=0, engine='openpyxl')
            elif file_extension == '.csv':
                return pd.read_csv(BytesIO(file_content))
            else:
                raise ValueError(f"Unsupported file format: {file_extension}")
        except Exception as e:
            logger.error(f"Error reading file: {str(e)}")
            raise

    @staticmethod
    def read_multi_sheet_excel(file_content: bytes) -> Dict[str, pd.DataFrame]:
        """Read all sheets from Excel file"""
        try:
            excel_file = pd.ExcelFile(BytesIO(file_content), engine='openpyxl')
            sheets = {}
            for sheet_name in excel_file.sheet_names:
                sheets[sheet_name] = pd.read_excel(excel_file, sheet_name=sheet_name)
            return sheets
        except Exception as e:
            logger.error(f"Error reading multi-sheet Excel: {str(e)}")
            raise

    @staticmethod
    def write_csv(data: List[Dict[str, Any]], columns: List[str] = None) -> str:
        """Write data to CSV string"""
        try:
            if not data:
                return ""
            
            df = pd.DataFrame(data)
            
            if columns:
                existing_columns = [col for col in columns if col in df.columns]
                df = df[existing_columns]
            
            df = df.fillna('')
            
            return df.to_csv(index=False)
        except Exception as e:
            logger.error(f"Error writing CSV: {str(e)}")
            raise

    @staticmethod
    def write_excel(data: List[Dict[str, Any]], columns: List[str] = None) -> bytes:
        """Write data to Excel bytes with formatting"""
        try:
            output = BytesIO()
            
            if not data:
                workbook = xlsxwriter.Workbook(output, {'in_memory': True})
                worksheet = workbook.add_worksheet()
                workbook.close()
                return output.getvalue()
            
            df = pd.DataFrame(data)
            
            if columns:
                existing_columns = [col for col in columns if col in df.columns]
                df = df[existing_columns]
            
            df = df.fillna('')
            
            workbook = xlsxwriter.Workbook(output, {'in_memory': True})
            worksheet = workbook.add_worksheet()
            
            header_format = workbook.add_format({
                'bold': True,
                'bg_color': '#4F81BD',
                'font_color': 'white',
                'border': 1,
                'align': 'center',
                'valign': 'vcenter'
            })
            
            cell_format = workbook.add_format({
                'border': 1,
                'valign': 'top',
                'text_wrap': True
            })
            
            for col_num, column in enumerate(df.columns):
                worksheet.write(0, col_num, column, header_format)
                
                column_width = max(len(str(column)), 
                                  df[column].astype(str).str.len().max() if not df.empty else 0)
                worksheet.set_column(col_num, col_num, min(column_width + 2, 50))
            
            for row_num, row_data in enumerate(df.values, start=1):
                for col_num, cell_value in enumerate(row_data):
                    worksheet.write(row_num, col_num, cell_value, cell_format)
            
            worksheet.freeze_panes(1, 0)
            
            workbook.close()
            output.seek(0)
            
            return output.getvalue()
            
        except Exception as e:
            logger.error(f"Error writing Excel: {str(e)}")
            raise

    @staticmethod
    def write_multi_sheet_excel(sheets_data: Dict[str, List[Dict[str, Any]]]) -> bytes:
        """Write multiple sheets to Excel file"""
        try:
            output = BytesIO()
            workbook = xlsxwriter.Workbook(output, {'in_memory': True})
            
            header_format = workbook.add_format({
                'bold': True,
                'bg_color': '#4F81BD',
                'font_color': 'white',
                'border': 1,
                'align': 'center',
                'valign': 'vcenter'
            })
            
            cell_format = workbook.add_format({
                'border': 1,
                'valign': 'top',
                'text_wrap': True
            })
            
            for sheet_name, data in sheets_data.items():
                worksheet = workbook.add_worksheet(sheet_name[:31])
                
                if not data:
                    continue
                
                df = pd.DataFrame(data)
                df = df.fillna('')
                
                for col_num, column in enumerate(df.columns):
                    worksheet.write(0, col_num, column, header_format)
                    
                    column_width = max(len(str(column)), 
                                      df[column].astype(str).str.len().max() if not df.empty else 0)
                    worksheet.set_column(col_num, col_num, min(column_width + 2, 50))
                
                for row_num, row_data in enumerate(df.values, start=1):
                    for col_num, cell_value in enumerate(row_data):
                        worksheet.write(row_num, col_num, cell_value, cell_format)
                
                worksheet.freeze_panes(1, 0)
            
            workbook.close()
            output.seek(0)
            
            return output.getvalue()
            
        except Exception as e:
            logger.error(f"Error writing multi-sheet Excel: {str(e)}")
            raise

    @staticmethod
    def get_column_mapping(entity_type: str) -> Dict[str, str]:
        """Get column mapping for different entity types"""
        
        mappings = {
            'products': {
                'Handle': 'handle',
                'Title': 'title',
                'Body (HTML)': 'body_html',
                'Vendor': 'vendor',
                'Product Type': 'product_type',
                'Tags': 'tags',
                'Published': 'published',
                'Option1 Name': 'option1',
                'Option1 Value': 'option1_value',
                'Option2 Name': 'option2',
                'Option2 Value': 'option2_value',
                'Option3 Name': 'option3',
                'Option3 Value': 'option3_value',
                'Variant SKU': 'sku',
                'Variant Grams': 'grams',
                'Variant Inventory Tracker': 'inventory_management',
                'Variant Inventory Qty': 'inventory_quantity',
                'Variant Inventory Policy': 'inventory_policy',
                'Variant Fulfillment Service': 'fulfillment_service',
                'Variant Price': 'price',
                'Variant Compare At Price': 'compare_at_price',
                'Variant Requires Shipping': 'requires_shipping',
                'Variant Taxable': 'taxable',
                'Variant Barcode': 'barcode',
                'Image Src': 'image_src',
                'Image Position': 'image_position',
                'Image Alt Text': 'image_alt',
                'SEO Title': 'metafield_global_title_tag',
                'SEO Description': 'metafield_global_description_tag',
                'Status': 'status',
                'Command': 'command'
            },
            'customers': {
                'First Name': 'first_name',
                'Last Name': 'last_name',
                'Email': 'email',
                'Company': 'company',
                'Address1': 'address1',
                'Address2': 'address2',
                'City': 'city',
                'Province': 'province',
                'Province Code': 'province_code',
                'Country': 'country',
                'Country Code': 'country_code',
                'Zip': 'zip',
                'Phone': 'phone',
                'Accepts Marketing': 'accepts_marketing',
                'Total Spent': 'total_spent',
                'Total Orders': 'orders_count',
                'Tags': 'tags',
                'Note': 'note',
                'Tax Exempt': 'tax_exempt',
                'Command': 'command'
            },
            'orders': {
                'Name': 'name',
                'Email': 'email',
                'Financial Status': 'financial_status',
                'Fulfillment Status': 'fulfillment_status',
                'Currency': 'currency',
                'Subtotal': 'subtotal_price',
                'Shipping': 'total_shipping_price',
                'Taxes': 'total_tax',
                'Total': 'total_price',
                'Paid at': 'paid_at',
                'Created at': 'created_at',
                'Lineitem quantity': 'line_items',
                'Lineitem name': 'line_item_name',
                'Lineitem price': 'line_item_price',
                'Billing Name': 'billing_name',
                'Billing Street': 'billing_street',
                'Billing City': 'billing_city',
                'Billing Province': 'billing_province',
                'Billing Country': 'billing_country',
                'Billing Zip': 'billing_zip',
                'Shipping Name': 'shipping_name',
                'Shipping Street': 'shipping_street',
                'Shipping City': 'shipping_city',
                'Shipping Province': 'shipping_province',
                'Shipping Country': 'shipping_country',
                'Shipping Zip': 'shipping_zip',
                'Notes': 'note',
                'Tags': 'tags'
            }
        }
        
        return mappings.get(entity_type, {})

    @staticmethod
    def normalize_dataframe(df: pd.DataFrame, entity_type: str) -> pd.DataFrame:
        """Normalize DataFrame column names based on entity type"""
        mapping = FileProcessor.get_column_mapping(entity_type)
        
        if mapping:
            reverse_mapping = {v: k for k, v in mapping.items()}
            df = df.rename(columns=reverse_mapping)
        
        return df

    @staticmethod
    def extract_metafields(row: Dict[str, Any]) -> List[Dict[str, Any]]:
        """Extract metafields from row data"""
        metafields = []
        
        for key, value in row.items():
            if key.startswith('Metafield:'):
                parts = key.replace('Metafield:', '').split('[')
                if len(parts) == 2:
                    namespace = parts[0].strip()
                    meta_key = parts[1].replace(']', '').strip()
                    
                    metafields.append({
                        'namespace': namespace,
                        'key': meta_key,
                        'value': value,
                        'type': 'string'
                    })
        
        return metafields
